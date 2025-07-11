import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook

#Año evaluado
Añoeval = 2025

#Funcion para definir numero del mes segun nombre
def mes_a_numero():
  return {
    "Enero": 1,
    "Febrero": 2,
    "Marzo": 3,
    "Abril": 4,
    "Mayo": 5,
    "Junio": 6,
    "Julio": 7,
    "Agosto": 8,
    "Septiembre": 9,
    "Octubre": 10,
    "Noviembre": 11,
    "Diciembre": 12
  }

#Funcion para convertir mes de ingles a español
def mes_eng_esp():
  return {
    "January":"Enero",
    "February":"Febrero",
    "March":"Marzo",
    "April":"Abril",
    "May":"Mayo",
    "June":"Junio",
    "July":"Julio",
    "August":"Agosto",
    "September":"Septiembre",
    "October":"Octubre",
    "November":"Noviembre",
    "December":"Diciembre"
  }

#Funcion para leer los archivos de la COOISPI
def leer_archivo(ruta, hoja=None):
  try:
    if not os.path.exists(ruta):  #Validar si la ruta existe
      raise FileNotFoundError(f"La ruta '{ruta}' no existe.")
    print("Leyendo archivo...")
    if ruta.lower().endswith('.csv'):  #Si la ruta es un csv
      df = pd.read_csv(ruta)
      return df
    elif ruta.lower().endswith(('.xlsx', 'xls')):  #Sino, si es un xlsx o xls
      excel = pd.ExcelFile(ruta)

      if hoja not in excel.sheet_names:  #Verificar que la hoja solicitada exista
        raise ValueError(f"La hoja '{hoja}' no existe en el archivo. Hojas disponibles: {excel.sheet_names}")

      df = pd.read_excel(excel, sheet_name=hoja)  #Leer directamente la hoja
      return df
    else:
      raise ValueError("Formato no soportado. Solo se permiten archivos CSV y Excel.")
  except Exception as e:
    raise RuntimeError(f"Error al leer archivo: {e}")

#Funcion para renombrar las columnas de la COOISPI
def renombrar_cooispi(df):

  #Limpiar columnas(eliminar espacios al inicio y al final)
  df.columns = [col.strip() for col in df.columns]

  #Definir posibles diccionarios de renombrado
  opciones_renombrado = [
    {
      'Unnamed: 0': 'Orden',
      'Unnamed: 1': 'Material',
      'Unnamed: 2': 'Texto de material',
      'Unnamed: 3': 'Ctd.',
      'Unnamed: 4': 'Lote',
      'Unnamed: 5': 'Almacén',
      'Unnamed: 6': 'Cl.mov.',
      'Unnamed: 7': 'Unidad',
      'Unnamed: 8': 'Doc.mat.',
      'Unnamed: 9': 'Pto.desc.',
      'Unnamed: 10': 'Fecha',
      'Unnamed: 11': 'ctd.UME',
      'Unnamed: 12': 'Impte.ML'
    },
    {
      'Orden': 'Orden',
      'Material': 'Material',
      'Texto de material': 'Texto de material',
      'Ctd.en UM base y signo +/- (MEINS)': 'Ctd.',
      'Lote': 'Lote',
      'Almacén': 'Almacén',
      'Clase de movimiento': 'Cl.mov.',
      'Unidad medida base (=MEINS)': 'Unidad',
      'Documento material': 'Doc.mat.',
      'Puesto descarga': 'Pto.desc.',
      'Fe.contabilización': 'Fecha',
      'Ctd.en UM entrada (/CWM/ERFME)': 'ctd.UME',
      'Importe ML (WAERS)': 'Impte.ML'
    }
  ]

  df = df.copy(deep=True) #Crear una copia completa de todo df

  columnas_actuales = set(df.columns)  #Identificar las columnas actuales del DataFrame

  #Seleccionar el diccionario de renombrado adecuado
  renombrado = False    #variable creada ya que al entrar al ciclo si solo una de las opciones de renombre no funciona(y hay dos) muestra el mensaje "No se encontraron columnas..."
  for opcion in opciones_renombrado:
    if set(opcion.keys()).issubset(columnas_actuales):
      df.rename(columns=opcion, inplace=True)
      renombrado = True   #se renombra y la variable pasa a True
      break

  #Si no coincidió con ningun diccionario, no se renombró y ahora sí se muestra el mensaje
  if not renombrado: 
    print("No se encontraron columnas coincidentes para renombrar.")

  return df

#Funcion para normalizar dataframes
def normalizar_dataframe(df):
  if df.empty:
    print("El DataFrame está vacío. No es posible normalizar.")
    return df
  for col in df.columns:
    #Convertir a str y eliminar espacios en columnas de tipo 'object'/strings
    if df[col].dtype == 'object':
      df[col] = df[col].astype(str).str.strip()
  return df

#Funcion para limpiar y unir dos DataFrames
def procesar_cooispi(df1, df2):
  if df1.empty or df2.empty:
    raise ValueError(f"Uno o ambos DataFrame están vacíos. Se requieren ambos para el procesamiento.")

  df2 = df2.astype(df1.dtypes)  #Cambia los tipos de datos de las columnas de 2025 para que coincidan con los de 2024

  # Aplicar normalización(limpieza de columnas) a ambos DataFrames
  df1 = normalizar_dataframe(df1)
  df2 = normalizar_dataframe(df2)

  if list(df1.columns) != list(df2.columns):
    raise ValueError(f"Las columnas de {df1} y {df2} no coinciden. No se puede continuar.")

  df_COOISPI = pd.concat([df1, df2], ignore_index=True)  #Une ambos DataFrames verticalmente

  df_COOISPI = df_COOISPI.copy(deep=True)  #Crea una copia completa de df_COOISPI

  return df_COOISPI

#Funcion para eliminar filas donde el valor de 'Numero'/columna este vacío (None o NaN)
def eliminar_filas_vacias(df, columna):
  filtrado = df[(df[columna].isnull()) | (df[columna] == "")]
  return filtrado

#Funcion para crear la columna Fecha y IdMes
def crear_fecha(df):
  df = df.copy()
  columnas_necesarias = ['Año:', 'Mes:', 'Dia']
  for col in columnas_necesarias:
    if col not in df.columns:
      print(f"La columna '{col}' no existe en el DataFrame.")
  df['IdMes'] = df['Mes:'].map(mes_a_numero())
  df['Fecha'] = df['Año:'].astype(str) + '-' + df['IdMes'].astype(str).str.zfill(2) + '-' + df['Dia'].astype(str).str.zfill(2)
  return df

#Funcion para tomar los datos desde el 2024 en adelante
def filtrar_desde_anio(df, año):
  df = df.copy()
  df['Año:'] = df['Año:'].astype(int)
  df = df[df['Año:'] >= año]
  return df

#Funcion para eliminar valores nulos, ceros y vacíos
def filtrar_nulos(df, columnas):
  filtro_nan_ceros_vacios = (
    (df[columnas].isna()).any(axis=1) |  #Filtra valores NaN
    (df[columnas] == 0).any(axis=1) |    #Filtra valores igual a 0
    (df[columnas] == "").any(axis=1)     #Filtra valores vacíos
  )
  df_nan_ceros = df[filtro_nan_ceros_vacios]   #Filas con errores
  df = df[~filtro_nan_ceros_vacios]            #Filas buenas
  return df_nan_ceros, df

#Funcion para validar que Peso promedio por unidad y Gramaje estén correctas
def validar_numericos(df, columna):
  if columna not in df.columns:
    print(f"La columna '{columna}' no existe en el DataFrame.")
    return df
  df = df.copy()
  df[columna] = df[columna].astype(str)                                  #Paso 1: Convertir todos los valores de la columna a strings
  df[columna] = df[columna].str.replace(',', '.').str.replace(';', '.')  #Paso 2: Reemplazar comas por puntos en caso de que se usen comas como separador decimal
  df[columna] = pd.to_numeric(df[columna], errors='coerce')              #Paso 3: Convertir los valores a float, forzando la conversión y estableciendo errores como NaN

  #Ver los valores no convertibles
  df_non_conver = df[df[columna].isna()]
  df = df[~df[columna].isna()]              #filtra los que NO son nan en df_Emp_2
  return df_non_conver, df

#Crea una columna Codigo con los 7 primeros caracteres de Código y Descrip. / Producto:
def extraer_codigo(df):
  if 'Código y Descrip. / Producto:' not in df.columns:
    print(f"La columna 'Código y Descrip. / Producto:' no existe en el DataFrame.")
    return df
  df['Codigo'] = df['Código y Descrip. / Producto:'].str.slice(0,7)
  return df

#Funcion para la identificacion y supresion de codigos con valores nulos
def eliminar_codigos_nan(df):
  df_Codigo_nan = df[df['Codigo'].isna()]
  indices_to_drop_cod_nan = df[df['Codigo'].isna()].index
  df_filtrado = df.drop(indices_to_drop_cod_nan)
  return df_filtrado, df_Codigo_nan

#Funcion para calcular el Real empacado, Deber ser, Diferencia y Sorbrepeso del archivo original de TPM
def calcular_columnas(df):
  df = df.copy()
  df['Real_empaque_calculado'] = df['Unidades Producidas (Conformes) :'] * df['Peso Promedio de la unidad (K):']
  df['Debe_ser_empaque_calculado'] = df['Gramaje (K):'] * df['Unidades Producidas (Conformes) :']
  df['Diferencia_calculado'] = df['Real_empaque_calculado'] - df['Debe_ser_empaque_calculado']
  df['Sobrepeso_calculado'] = df['Diferencia_calculado'] / df['Debe_ser_empaque_calculado']
  return df

#Funcion para la generacion de novedades mayores al 50%
def generar_novedades(df, Sobrepeso_Novedades):
  df_Nov_Sobrepeso = df[(df['Sobrepeso_calculado'] >= Sobrepeso_Novedades) & (df['Año:'] == Añoeval)]   #Filtra los registros del añoeval con sobrepeso >= al límite
  return df_Nov_Sobrepeso

#Funcion para crear tabla pivote para le consolidado mensual
def pivote_consolidado(df):
  if df.empty:
    print("El DataFrame está vacío.")
    return df
  df_Mes = df.pivot_table(
    index=['Año:', 'Mes:','Codigo','Máquina / Equipo:'],
    values=['Unidades Producidas (Conformes) :','Gramaje (K):','Peso Promedio de la unidad (K):','Real_empaque_calculado','Debe_ser_empaque_calculado','Diferencia_calculado'],
    aggfunc={'Unidades Producidas (Conformes) :': 'sum','Gramaje (K):':'mean',
              'Peso Promedio de la unidad (K):':'mean',
              'Real_empaque_calculado':'sum',
              'Debe_ser_empaque_calculado':'sum',
              'Diferencia_calculado':'sum'}
  ).reset_index()
  return df_Mes

#Funcion para estandarizar meses en ingles
def estandarizar_mes(df):
  if 'Fecha' not in df.columns:
    print(f"La columna 'Fecha' no existe en el DataFrame.")
    return df
  df = df.copy()
  df['Mes:'] = df['Fecha'].dt.strftime('%B').str.capitalize().map(mes_eng_esp())
  df['Año:'] = df['Fecha'].dt.year.astype(int)
  return df

#Funcion para filtrar columna por valor especifico
def filtrar_columna(df, columna, valores):
  if columna not in df.columns:
    print(f"La columna '{columna}' no existe en el DataFrame.")
    return df
  return df[df[columna].isin(valores)]

#Funcion para quitar los Materiales que empiezan en 8
def filtrar_material(df, prefijo):
  if 'Material' not in df.columns:
    print("La columna 'Material' no existe en el DataFrame.")
    return df
  return df[~df['Material'].astype(str).str.startswith(str(prefijo))]

#Funcion para realizar un "left join" entre df_COOISPI_261_262 y df_COOISPI_101_102
def merge_cooispi(df2, df1):
  df_COOISPI_COMB = pd.merge(
    df2,                                              #DataFrame base
    df1[['Orden', 'Material', 'Texto de material']],  #DataFrame a unir
    on='Orden',                                       #Llave de unión
    how='left',                                       #Se quedan todas las filas de la izquierda(df2) y solo las coincidentes de la derecha(df1)
    suffixes=('', '_producto_term')                   #Si hay columnas con el mismo nombre, las del segundo DataFrame tendrán el sufijo _producto_term
  )
  #Eliminar filas duplicadas
  df_COOISPI_COMB.drop_duplicates(inplace=True)
  return df_COOISPI_COMB

#Funcion para la creacion de la agrupacion mensual del archivo semielaborado
def pivote_semielaborados(df):
  if df.empty:
    print("El DataFrame está vacío.")
    return df
  df_costo_semi_Mes = df.pivot_table(
    index=['Año:', 'Mes:', 'Codigo'],
    values=['Ctd.', 'Impte.ML'],
    aggfunc='sum'
  ).reset_index()

  df_costo_semi_Mes['Costo/kg'] = df_costo_semi_Mes['Impte.ML'] / df_costo_semi_Mes['Ctd.']
  return df_costo_semi_Mes

#Funcion para traer la columna 'Costo/kg' al dataframe original
def merge_costo(df_2, df_costo_semi_Mes):
  df_2 = df_2.merge(
    df_costo_semi_Mes[['Año:', 'Mes:', 'Codigo', 'Costo/kg']],
    on=['Año:', 'Mes:', 'Codigo'],
    how='left'
  )
  return df_2

#Funcion para eliminar la columna Costo/kg antigua y renombrar la nueva
def modificar_costo(df_2):
  df_2.drop(columns=['Costo/kg_x'], inplace=True)
  df_2.rename(columns={'Costo/kg_y': 'Costo/kg'}, inplace=True)
  return df_2

#Calculo estandar para hallar los Ahorros_Perdidas
def calcular_ahorros_perdidas(df):
  df['Ahorros/Perdidas'] = (df['Meta'] - df['Sobrepeso_calculado']) * df['Real_empaque_calculado'] * df['Costo/kg']
  return df

#Funcion para mostrra si hay o no filas repetidas en un dataframe
def filas_repetidas(df):
  duplicated_rows = df[df.duplicated()]
  print(f"Filas repetidas: '{duplicated_rows}'" if not duplicated_rows.empty else "No hay filas duplicadas.")

#Fusionar los DataFrames en base a 'Año:', 'Codigo' y 'Mes:'
def unir_dataframes(df_totalizado_Mes, df_costo_semi_Mes):
  if df_totalizado_Mes.empty or df_costo_semi_Mes.empty:
    print("Uno o ambos Dataframes están vacíos.")
    return df_totalizado_Mes

  #Fusionar usando merge(JOIN en SQL)
  df_totalizado_Mes = pd.merge(
    df_totalizado_Mes,
    df_costo_semi_Mes[['Año:', 'Codigo', 'Mes:', 'Ctd.', 'Impte.ML']],
    on=['Año:', 'Codigo', 'Mes:'],
    how='left'
  )

  #Calcular el promedio de 'Ctd.' y 'Impte.ML' por cada Codigo
  promedios = df_totalizado_Mes.groupby('Codigo')[['Ctd.', 'Impte.ML']].transform('mean')

  #Llenar los valores NaN con el promedio correspondiente al Codigo
  df_totalizado_Mes[['Ctd.', 'Impte.ML']] = df_totalizado_Mes[['Ctd.', 'Impte.ML']].fillna(promedios)

  return df_totalizado_Mes

#Funcion para calcular las columnas Real_empaque, Debe_ser_empaque, Diferencia, Sobrepeso, Meta, Costos y Ahorros_Perdidas
def calcular_columnas_totalizado(df, meta):
  df['Real_empaque'] = df['Real_empaque_calculado']
  df['Debe_ser_empaque'] = df['Debe_ser_empaque_calculado']
  df['Diferencia'] = df['Diferencia_calculado']
  df['Sobrepeso'] = df['Diferencia'] / df['Debe_ser_empaque']
  df['Meta'] = meta
  df['Costo/kg'] = df['Impte.ML'] / df['Ctd.']
  df['Ahorros_Perdidas'] = (df['Meta'] - df['Sobrepeso']) * df['Real_empaque'] * df['Costo/kg']
  return df

#Funcion para calcular las columnas Real_empaque, Debe_ser_empaque, Diferencia, Sobrepeso y Meta para mezclas segun maquina
def calcular_columnas_totalizado_mezclas(df):
  df['Real_empaque'] = df['Unidades Producidas (Conformes) :'] * (df['Peso Promedio de la unidad (K):']/1000)
  df['Debe_ser_empaque'] = df['Unidades Producidas (Conformes) :'] * (df['Gramaje (K):']/1000)
  df['Diferencia'] = df['Real_empaque'] - df['Debe_ser_empaque']
  df['Sobrepeso'] = df['Diferencia'] / df['Debe_ser_empaque']
  df['Meta'] = df['Máquina / Equipo:'].map({'ENFLEX': 0.0056,
                                            'TRANSPACK 2': 0.0033,
                                            'INGEPACK': 0.0072,
                                            'ROVEMA':0.0078,
                                            'LANIC':0.0055,
                                            'TOYO 7':0.0065,
                                            'TOYO':0.0088,
                                            'ENCAPSULADORA':0.006,
                                            'INGEPACK 2':0.0068})

  df['Costo/kg'] = df['Impte.ML'] / df['Ctd.']
  df['Ahorros_Perdidas'] = (df['Meta'] - df['Sobrepeso']) * df['Real_empaque'] * df['Costo/kg']
  return df

#Funcion para traer los codigos sin duplicados, con su codigo/descripcion y su maquina
def obtener_cod_descrip(df_costo_semi, df_mes):
  df_codigos = df_costo_semi[['Codigo', 'TEXTO PT']].rename(columns={'TEXTO PT': 'Código y Descrip. / Producto:'})
  df = df_mes[['Codigo', 'Máquina / Equipo:']]
  df_Codigos = pd.merge(df_codigos, df, on='Codigo')
  df_Codigos = df_Codigos.drop_duplicates()
  return df_Codigos

#Funcion para eliminar duplicados por columna
def eliminar_duplicados_columna(df, nombre_columna):
  return df[[nombre_columna]].drop_duplicates().dropna(subset=[nombre_columna])

#Funcion para generacion de archivo de fechas automaticas
def generar_fechas(fecha_inicio):
  Fecha_Inicio = fecha_inicio  #Fecha de inicio

  Fecha_Final = datetime.today().strftime("%Y-%m-%d")  #Fecha actual

  Fechas = pd.date_range(start=Fecha_Inicio, end=Fecha_Final)  #Crear rango de fechas

  df_Fechas = pd.DataFrame(Fechas, columns=['Fecha'])  #Crear el dataframe de fechas

  df_Fechas.tail()  #Mostrar el dataframe

  return df_Fechas

#Funcion para actualizar hojas de excel
def actualizar_hojas_excel(file_path, dataframes):
  print("Procesando Consolidado...")
  # Verificar si el archivo ya existe
  try:
    book = load_workbook(file_path)  # Cargar el archivo existente
  except FileNotFoundError:
    book = None  # Si no existe, se creará uno nuevo

  # Abrir en modo de escritura
  with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    for sheet_name, df in dataframes.items():
      if df is not None and not df.empty:
        if book and sheet_name in book.sheetnames:
          # Limpiar el contenido de la hoja antes de escribir
          sheet = book[sheet_name]
          for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
              cell.value = None  # Borra el contenido (excepto los encabezados)
          book.save(file_path)  # Guardar cambios en el archivo

        # Escribir nuevos datos en la hoja limpia
        df.to_excel(writer, sheet_name=sheet_name, index=False)
      else:
        print(f"⚠️ Advertencia: {sheet_name} está vacío y no se escribirá.")
  print(f"✅ Las hojas en el Archivo de Consolidado han sido limpiadas y actualizadas correctamente.")

#Funcion para guardar el archivo de novedades con varias hojas
def crear_archivo_novedades(ruta, dataframes):
  print("Procesando Novedades...")
  hojas_guardadas = 0

  with pd.ExcelWriter(ruta, engine='xlsxwriter') as writer:
    for nombre_hoja, df in dataframes.items():
      if df is not None:
        df.to_excel(writer, sheet_name=nombre_hoja, index=False)
        hojas_guardadas += 1  #Contamos cuantas hojas se han guardado

  #Verificamos si el número de hojas guardadas es igual al número de hojas que queríamos guardar
  if hojas_guardadas == len(dataframes):
    print(f"✅ El Archivo de Novedades guardado exitosamente con {hojas_guardadas} hojas guardadas correctamente.")
  else:
    print(f"⚠️ Advertencia: No todas las hojas fueron guardadas correctamente.")

####################################################################################################################


# wb = load_workbook(Ruta_Consolidados)
# print("Hojas visibles en el archivo:")
# for sheet in wb.sheetnames:
#     estado = wb[sheet].sheet_state
#     print(f" - {sheet} (estado: {estado})")

# #Eliminar hojas especificas de un libro en excel
# # Cargar el libro de Excel
# workbook = openpyxl.load_workbook(Ruta_Consolidados)
# # Eliminar la hoja "Hoja9"
# try:
#     hoja_a_eliminar = workbook["Hoja7"]
#     workbook.remove(hoja_a_eliminar)
# except KeyError:
#     print(f"La hoja '{hoja_a_eliminar}' no existe.")
# # Guardar el libro modificado
# workbook.save(Ruta_Consolidados)
